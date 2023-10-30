import re
import pandas as pd
import openpyxl

##realizar caso por caso, limpiar un caso y despues el otro
sustituciones = {
    r'\bCARRERA\b': 'CR',
    r'\bCALLE\b': 'CL',
    r'\bCIRCULAR\b': 'CQ',
    r'\bTRANSVERSAL\b': 'TV',
    r'\bDIAGONAL\b': 'DG',
}

palabras_clave = {"CARRERA", "CALLE", "CIRCULAR", "TRANSVERSAL", "DIAGONAL", "CL", "CR"}
borrar_palabras = {"CARRERA", "CALLE", "CIRCULAR", "TRANSVERSAL", "DIAGONAL", "CL", "CR","CARRERAS","CALLES","CARREREA"}
eliminar_caracteres = {";",",","<",">","/","&"}

def reemplazo(match):
    return sustituciones[match.group()]

def borrar_repetido(direcciones):
    direccion_limpia = []
    for direccion in direcciones:
        # Dividimos la cadena en palabras
        palabras = direccion.split()

        # Contadores para las coincidencias
        coincidencias = 0

        # Lista para almacenar las palabras válidas
        palabras_validas = []

        for palabra in palabras:
            if palabra in ["CL", "CR", "TV"]:
                coincidencias += 1
                if coincidencias == 2:
                    break  # Detenemos el proceso si encontramos la segunda coincidencia
            palabras_validas.append(palabra)
            # Ahora puedes unir las palabras válidas en una cadena nuevamente
        resultado = ' '.join(palabras_validas)
        direccion_limpia.append(resultado)
    return direccion_limpia

def eliminar_palabras(direcciones):
    direccion_limpia =[]
    for direccion in direcciones:
        # Definir un patrón de expresión regular que coincide con "CR," "CL," "TV," "CQ," "SUR," "NORTE", combinaciones alfanuméricas como "80A", "10E" o "5D", y también con combinaciones numéricas como "123"
        #patron = r'\b(?:CR|CL|TV|CQ|SUR|NORTE|\d{2,3}[A-Z]|\d{1}[A-Z]|\d{1,3})\b'
        #patron = r'\b(?:CR|CL|TV|CQ|AV|SUR|NORTE|\d{1,3}(?:[A-Z]{1,2})?)\b'
        #patron = r'\b(?:CR|CL|TV|CQ|AV|SUR|NORTE|\d{1,3}(?:[A-Z]{1,2}|\d{1,4}[A-Z]{1,4})?)\b'
        patron = r'\b(?:CR|CL|TV|CQ|AV|SUR|NORTE|\d{1,3}(?:[A-Z]{1,4})?)\b'



        # Dividir la dirección en palabras
        palabras = direccion.split()

        # Inicializar una lista vacía para las palabras que cumplen con el patrón
        palabras_filtradas = []

        # Recorrer cada palabra y verificar si cumple con el patrón
        encontrado = False
        for palabra in palabras:
            
            if re.search(patron, palabra) and encontrado == False:
                palabras_filtradas.append(palabra)
            else:
                encontrado = True
        direccion_estandarizada = ' '.join(palabras_filtradas)
        direccion_limpia.append(direccion_estandarizada)
    return direccion_limpia

def borrar_union(direcciones):
    direccion_limpia=[]
    # Expresión regular para encontrar "CALLE" o "CARRERA" y capturar todo antes de ellos
    pattern = re.compile(r'^(.*?)(CALLE|CARRERA|TORRE|PARQUEADERO)')

    for direccion in direcciones:
        match = re.search(pattern, direccion)
        if match:
            inicio_direccion = match.group(1)
            direccion_limpia.append(inicio_direccion)
        else:
            direccion_limpia.append(direccion)
    return direccion_limpia

def borrar_entre(direcciones):
    direcciones_estandarizadas = []

    for direccion in direcciones:
        palabras = direccion.split()
        if 'ENTRE' in palabras or 'POR' in palabras or 'CON' in palabras:
            if 'ENTRE' in palabras:
                indice_entre = palabras.index('ENTRE')
                if indice_entre > 0 and palabras[indice_entre + 1] in borrar_palabras:
                    palabras[indice_entre:indice_entre + 2] = ['']
                    if 'Y' in palabras[indice_entre:]:
                        indice_y = palabras.index('Y', indice_entre)
                        palabras = palabras[:indice_y]
                    if 'Y' in palabras:
                        palabras = [palabra for palabra in palabras if palabra != 'Y']
                        
            if 'POR' in palabras:
                indice_entre = palabras.index('POR')
                if indice_entre > 0 and palabras[indice_entre + 1] in borrar_palabras:
                    palabras[indice_entre:indice_entre + 2] = ['']
                    if 'Y' in palabras[indice_entre:]:
                        indice_y = palabras.index('Y', indice_entre)
                        palabras = palabras[:indice_y]
                    if 'Y' in palabras:
                        palabras = [palabra for palabra in palabras if palabra != 'Y']
            if 'CON' in palabras:
                indice_entre = palabras.index('CON')
                if indice_entre > 0 and palabras[indice_entre + 1] in borrar_palabras:
                    palabras[indice_entre:indice_entre + 2] = ['']
                    if 'Y' in palabras[indice_entre:]:
                        indice_y = palabras.index('Y', indice_entre)
                        palabras = palabras[:indice_y]
                    if 'Y' in palabras:
                        palabras = [palabra for palabra in palabras if palabra != 'Y']

            direccion_final = ' '.join(palabra for palabra in palabras if palabra)
            direcciones_estandarizadas.append(direccion_final.strip())
        else:
            direccion_final = ' '.join(palabra for palabra in palabras if palabra)
            direcciones_estandarizadas.append(direccion_final.strip())
    return direcciones_estandarizadas
def estandarizar_direccion(direcciones):
    nuevas_direcciones = []
    print(direcciones)
    nueva_cadena = []
    for direccion in direcciones:
        palabras = direccion.split()
        print(palabras)
        if palabras[0] == "CR":
            print(palabras[0])
            nueva_cadena.append(palabras[0])
            print
            if re.match(r'^[0-9]{1,3}$', palabras[1]):
                nueva_cadena.append(palabras[1])
            elif re.match(r'^[0-9]{1,3}[A-Za-z]{0,2}$', palabras[1]):
                nueva_cadena.append(palabras[1])
        direccion_unida = " ".join(nueva_cadena)
        nuevas_direcciones.append(direccion_unida)
    return nuevas_direcciones

def limpiar_direccion(direcciones):
    direccion_limpia = []
    direccion_corta = []
    direccion_split = []

    for direccion in direcciones:
        nueva_cadena = ""
        encontrado = False
        direccion = str.upper(direccion)
        #print(direccion)
        if direccion.strip() == "":
            direccion_corta.append(direccion)

        elif direccion.startswith("12"):
            #cortar el punto y coma hacia la derecha o vaciar con cadena vacia
            print(direccion)
        elif direccion =="NAN" or direccion=="NULL":
            direccion = ""
            direccion_corta.append(direccion)
        elif direccion.startswith("<"):
            contenido = direccion[1:-1]  # Elimina los signos "<" y ">"
            #print(contenido)
            split_contenido = contenido.split()
            #print(split_contenido)
            if split_contenido[0] in palabras_clave:
                #print(split_contenido)
                contenido = " ".join(split_contenido)
                if ";" in contenido:
                    parte_anterior = contenido.split(";", 1)[0]
                    #otro split 
                    # union = ",".join(parte_anterior)***no se le hace join
                    # parte_anterior2 = union.split(">", 1)[0]
                    #print(parte_anterior)
                    dire_upper = str.upper(parte_anterior)
                    direccion_split = dire_upper.split()
                    #print(direccion_split)
                    #print(direccion_split)
                    direccion_unida= ""
                    if direccion_split[0] in palabras_clave:
                        direccion_unida = " ".join(direccion_split)
                        #print(direccion_unida)
                        direccion_unida = re.compile(r'<[^>]+>')
                        cadena_sin_contenido = direccion_unida.sub('', dire_upper)
                        contenido_entre_parentesis = re.compile(r'\([^)]+\)')
                        cadena_sin_parentesis = contenido_entre_parentesis.sub('', cadena_sin_contenido)
                        #print(cadena_sin_parentesis)
                        for caracter in cadena_sin_parentesis:
                            if caracter in eliminar_caracteres:
                                encontrado = True
                            if not encontrado:
                                nueva_cadena += caracter
                        direccion_corta.append(nueva_cadena)
                else:
                    direccion_unida = re.compile(r'<[^>]+>')
                    cadena_sin_contenido = direccion_unida.sub('', contenido)
                    contenido_entre_parentesis = re.compile(r'\([^)]+\)')
                    cadena_sin_parentesis = contenido_entre_parentesis.sub('', cadena_sin_contenido)
                    #print(cadena_sin_parentesis)
                    for caracter in cadena_sin_parentesis:
                        if caracter in eliminar_caracteres:
                            encontrado = True
                        if not encontrado:
                            nueva_cadena += caracter
                    direccion_corta.append(nueva_cadena)    

            elif split_contenido[0] not in palabras_clave:
                contenido = " ".join(split_contenido)
                if ";" in contenido:
                    parte_anterior = contenido.split(";", 1)[1]
                    dire_upper = str.upper(parte_anterior)
                    direccion_split = dire_upper.split()
                    #print(direccion_split)
                    #print(direccion_split)
                    direccion_unida= ""
                    if direccion_split[0] in palabras_clave:
                        direccion_unida = " ".join(direccion_split)
                        #print(direccion_unida)
                        direccion_unida = re.compile(r'<[^>]+>')
                        cadena_sin_contenido = direccion_unida.sub('', dire_upper)
                        contenido_entre_parentesis = re.compile(r'\([^)]+\)')
                        cadena_sin_parentesis = contenido_entre_parentesis.sub('', cadena_sin_contenido)
                        #print(cadena_sin_parentesis)
                        for caracter in cadena_sin_parentesis:
                            if caracter in eliminar_caracteres:
                                encontrado = True
                            if not encontrado:
                                nueva_cadena += caracter
                        direccion_corta.append(nueva_cadena)
                    else :
                            contenido = ""
                            direccion_corta.append(contenido)  
                else :
                    contenido = ""
                    direccion_corta.append(contenido)   

        else:
            dire_upper = str.upper(direccion)
            direccion_split = dire_upper.split()
            direccion_unida= ""
            
            if direccion_split[0] in palabras_clave:
                direccion_unida = " ".join(direccion_split)
                direccion_sin_caracteres = re.compile(r'<[^>]+>')
                cadena_sin_contenido = direccion_sin_caracteres.sub('', dire_upper)
                contenido_entre_parentesis = re.compile(r'\([^)]+\)')

                cadena_sin_parentesis = contenido_entre_parentesis.sub('', cadena_sin_contenido)

                for caracter in cadena_sin_parentesis:
                    if caracter in eliminar_caracteres:
                        encontrado = True
                    if not encontrado:
                        nueva_cadena += caracter
                
                direccion_corta.append(nueva_cadena)
            else:
                direccion_unida = " ".join(direccion_split)
                direccion_sin_caracteres = re.compile(r'<[^>]+>')
                # si dentro de  <> encuentra la palabra carrera,calle,cl...etc no me lo borre

                # Sustituye el contenido entre '<' y '>' con una cadena vacía
                cadena_sin_contenido = direccion_sin_caracteres.sub('', dire_upper)
                contenido_entre_parentesis = re.compile(r'\([^)]+\)')

                # Sustituye el contenido entre paréntesis con una cadena vacía
                cadena_sin_parentesis = contenido_entre_parentesis.sub('', cadena_sin_contenido)
                

                #print(cadena_sin_parentesis)
                for caracter in cadena_sin_parentesis:
                    if caracter in eliminar_caracteres:
                        ##if caract < equivale al primer caracter ejecute otra funcion que verifica si la informacion contenida si es valida
                        encontrado = True
                    if not encontrado:
                        nueva_cadena += caracter
                direccion_corta.append(nueva_cadena)
            

    return direccion_corta
data = pd.read_excel('paq1.xlsx')

ids = data["id"].tolist()
registros = data["Direccion"].tolist()

lista_de_registros = []

for registro in registros:
    dire_parse = str(registro)
    dire_upper = str.upper(dire_parse)
    lista_de_registros.append(dire_upper)

direccion_estandarizada = limpiar_direccion(lista_de_registros)

direccion_modificada = []

for direccion in direccion_estandarizada:
    for patron, reemplazo in sustituciones.items():
        direccion = re.sub(patron, reemplazo, direccion)
    direccion_modificada.append(direccion)

#direccion_sin_entre = borrar_entre(direccion_modificada)

direccion_limpia = borrar_union(direccion_modificada)
direccion_sin_palabras = eliminar_palabras(direccion_limpia)
direccion_sin_repetir = borrar_repetido(direccion_sin_palabras)


wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(row=1, column=1).value = 'ID'  # Nombre de la columna 1
sheet.cell(row=1, column=2).value = 'Dirección Estandarizada' 
sheet.cell(row=1, column=3).value = 'Dirección Original' 


for i, dato in enumerate(direccion_sin_repetir):
    sheet.cell(row=i + 2, column=1).value = ids[i]
    sheet.cell(row=i + 2, column=2).value = dato
    sheet.cell(row=i + 2, column=3).value = registros[i]

wb.save("archivo_prueba.xlsx")