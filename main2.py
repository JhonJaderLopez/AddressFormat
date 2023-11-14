import re
import pandas as pd
import openpyxl

##realizar caso por caso, limpiar un caso y despues el otro
sustituciones = {
    r'\bCARRERA\b': 'CR',
    r'\bCALLE\b': 'CL',
    r'\bCIRCULAR\b': 'CQ',
    r'\bTRANSVERSAL\b': 'TV',
    r'\bDIAGONAL\b': 'DG',
    r'\bAVENIDA\b': 'AV',
}

palabras_clave = {"CARRERA", "CALLE", "CIRCULAR", "TRANSVERSAL", "DIAGONAL","AVENIDA","DG", "CL", "CR","TV","CQ","AV"}
borrar_palabras = {"CARRERA", "CALLE", "CIRCULAR", "TRANSVERSAL", "DIAGONAL","AVENIDA" ,"DG","CL", "CR","TV","CQ","AV","CARRERAS","CALLES","CARREREA"}
eliminar_caracteres = {";",",","<",">","/","&","-","ª"}
def reemplazo(match):
    return sustituciones[match.group()]

def vaciar_invalidas(direcciones):
    direccion_limpia = []
    for direccion in direcciones:
        if direccion in palabras_clave:
            direccion = ""
            direccion_limpia.append(direccion)
        else:
            direccion_limpia.append(direccion)
    return direccion_limpia

def contar_palabras(cadena):
    palabras = re.findall(r'\S+', cadena)
    return len(palabras)

def borrar_repetido(direcciones):
    direccion_corta = []
    for direccion in direcciones:
        cadenas_separadas = direccion.split(';')
        # Inicializar variables para realizar un seguimiento de la cadena más larga y su longitud
        cadena_mas_larga = ""
        longitud_mas_larga = 0
        # Iterar a través de las cadenas separadas
        for cadena in cadenas_separadas:
            num_palabras = contar_palabras(cadena)
            if num_palabras > longitud_mas_larga:
                longitud_mas_larga = num_palabras
                cadena_mas_larga = cadena
        direccion_corta.append(cadena_mas_larga)
    return direccion_corta

def procesar_cadena(cadena):
    coincidencias = re.findall(r'\bCR\b|\bCL\b|\bTV\b|\bCQ\b|\bAV\b|\bDG\b', cadena, re.IGNORECASE)
    if len(coincidencias) == 2:
        palabras = cadena.split()
        contador = 0
        nueva_cadena = ""
        for palabra in palabras:
            if re.match(r'\bCR\b|\bCL\b|\bTV\b|\bCQ\b|\bAV\b|\bDG\b', palabra, re.IGNORECASE):
                contador += 1
                if contador == 2:
                    continue
            nueva_cadena += palabra + " "
        return nueva_cadena.strip()
    return cadena

def borrar_repetido_separada(direcciones):
    direccion_limpia = []
    for direccion_larga in direcciones:
        direccion_corta = []
        direccion_separada = direccion_larga.split(";")
        for direccion in direccion_separada:

            if direccion.strip() == "":
                direccion_corta.append(direccion)
            else:
                # Dividimos la cadena en palabras
                palabras = direccion.split()
                if palabras[0] not in palabras_clave:
                    direccion= ""
                    direccion_corta.append(direccion)
                else:
                    #condicion para entrar a eliminar coincnidenca#
                    coincidencias = 0
                    # Lista para almacenar las palabras válidas
                    palabras_validas1 = []
                    palabras_validas2 = []
                    for palabra in palabras:
                        if palabra in borrar_palabras:
                            coincidencias += 1
                        if coincidencias == 1:
                            palabras_validas1.append(palabra)
                        if coincidencias == 2 :
                            palabras_validas2.append(palabra)
                        if coincidencias == 3:
                            break
                    if len(palabras_validas1) < 3 and len(palabras_validas2) <= 3:
                        cadena= ' '.join(palabras) 
                        nueva_cadena = procesar_cadena(cadena)
                        palabras = nueva_cadena.split()
                    coincidencias = 0
                    # Lista para almacenar las palabras válidas
                    palabras_validas1 = []
                    palabras_validas2 = []

                    for palabra in palabras:
                        if palabra in borrar_palabras:
                            coincidencias += 1
                        if coincidencias == 1:
                            palabras_validas1.append(palabra)
                        if coincidencias == 2 and len(palabras_validas1)<3:
                            palabras_validas2.append(palabra)
                            palabras_validas1 = []
                        if coincidencias == 3:
                            break

                    if len(palabras_validas1) > 0:
                        resultado = ' '.join(palabras_validas1)
                    elif len(palabras_validas2) > 0:
                        resultado = ' '.join(palabras_validas2)
                        #ayer estaba tratando de vaciar la ultima lista cuando coincidencias ees = 2, revisar

                        #tengo que tener en cuenta 2 casos: cuando la correcta es la segunda coincidencia y cuando es la primera
                        #porque la larga puede ser la 1 o la 2, porque sino puedo borrar la corecta o incluso dejar vacio 

                        # Ahora puedes unir las palabras válidas en una cadena nuevamente
                    direccion_corta.append(resultado)

        nueva_direccion= ';'.join(direccion_corta)        
        direccion_limpia.append(nueva_direccion)
    return direccion_limpia

def eliminar_palabras(direcciones):
    direccion_limpia =[]
    for direccion_larga in direcciones:
        direccion_corta = []
        direccion_separada= direccion_larga.split(";")
        for direccion in direccion_separada:
            patron = r'\b(?:CR|CL|TV|CQ|AV|DG|SUR|ESTE|\d{1,3}(?:[A-Z]{1,4})?)\b'
            # Dividir la dirección en palabras
            palabras = direccion.split()
            # Inicializar una lista vacía para las palabras que cumplen con el patrón
            palabras_filtradas = []
            # Recorrer cada palabra y verificar si cumple con el patrón
            encontrado = False
            for palabra in palabras:
                if re.search(patron, palabra) and encontrado == False:
                    palabras_filtradas.append(palabra)
                else:
                    encontrado= True
                    pass
            direccion_estandarizada = ' '.join(palabras_filtradas)
            direccion_corta.append(direccion_estandarizada)
        nueva_direccion= ';'.join(direccion_corta)        
        direccion_limpia.append(nueva_direccion)
    return direccion_limpia

def borrar_union(direcciones):
    direccion_limpia=[]
    # Expresión regular para encontrar "CALLE" o "CARRERA" y capturar todo antes de ellos
    pattern = re.compile(r'^(.*?)(CALLE|CARRERA|TORRE|PARQUEADERO|DIAGONAL|TRANSVERSAL|CIRCULAR|AVENIDA)')
    #pattern = re.compile(r'^(.*?)(?<!\s)(CALLE|CARRERA|TORRE|PARQUEADERO|DIAGONAL|TRANSVERSAL|CIRCULAR|AVENIDA)')
    for direccion_larga in direcciones:
        direccion_corta = []
        direccion_separada = direccion_larga.split(";")
        for direccion in direccion_separada:
            match = re.search(pattern, direccion)
            if match:
                inicio_direccion = match.group(1)
                direccion_corta.append(inicio_direccion)
            else:
                direccion_corta.append(direccion)
        nueva_direccion= ';'.join(direccion_corta)        
        direccion_limpia.append(nueva_direccion)
    return direccion_limpia

def borrar_entre(direcciones):
    direccion_limpia = []
    for direccion_larga in direcciones:
        direcciones_estandarizadas = []
        direccion_final = ""
        direccion_separada = direccion_larga.split(";")
        for direccion in direccion_separada:
            #palabras = ""
            palabras = direccion.split()
            if 'ENTRE' in palabras or 'POR' in palabras or 'CON' in palabras or 'CRUCERO' in palabras or 'HACIA' in palabras or 'FRENTE' in palabras:
                if 'ENTRE' in palabras:
                    indice_entre = palabras.index('ENTRE')
                    if indice_entre > 0 and (indice_entre + 1) < len(palabras) and palabras[indice_entre + 1] in borrar_palabras:
                        palabras[indice_entre:indice_entre + 2] = ['']
                        if 'Y' in palabras[indice_entre:]:
                            indice_y = palabras.index('Y', indice_entre)
                            palabras = palabras[:indice_y]
                        if 'Y' in palabras:
                            palabras = [palabra for palabra in palabras if palabra != 'Y']
                    elif indice_entre > 0 and (indice_entre + 2) < len(palabras) and palabras[indice_entre + 2] in borrar_palabras:
                        palabras[indice_entre:indice_entre + 3] = ['']
                        if 'Y' in palabras[indice_entre:]:
                            indice_y = palabras.index('Y', indice_entre)
                            palabras = palabras[:indice_y]
                        if 'Y' in palabras:
                            palabras = [palabra for palabra in palabras if palabra != 'Y']
                    elif indice_entre > 0 and (indice_entre) < len(palabras):
                        palabras[indice_entre:indice_entre + 1] = ['']
                        if 'Y' in palabras[indice_entre:]:
                            indice_y = palabras.index('Y', indice_entre)
                            palabras = palabras[:indice_y]
                        if 'Y' in palabras:
                            palabras = [palabra for palabra in palabras if palabra != 'Y']
                if 'POR' in palabras:
                    indice_entre = palabras.index('POR')
                    if indice_entre > 0 and (indice_entre + 1) < len(palabras) and palabras[indice_entre + 1] in borrar_palabras:
                        palabras[indice_entre:indice_entre + 2] = ['']
                        if 'Y' in palabras[indice_entre:]:
                            indice_y = palabras.index('Y', indice_entre)
                            palabras = palabras[:indice_y]
                        if 'Y' in palabras:
                            palabras = [palabra for palabra in palabras if palabra != 'Y']
                    elif indice_entre > 0 and (indice_entre + 2) < len(palabras) and palabras[indice_entre + 2] in borrar_palabras:
                        palabras[indice_entre:indice_entre + 3] = ['']
                        if 'Y' in palabras[indice_entre:]:
                            indice_y = palabras.index('Y', indice_entre)
                            palabras = palabras[:indice_y]
                        if 'Y' in palabras:
                            palabras = [palabra for palabra in palabras if palabra != 'Y']
                    elif indice_entre > 0 and (indice_entre + 3) < len(palabras) and palabras[indice_entre + 3] in borrar_palabras:
                        palabras[indice_entre:indice_entre + 4] = ['']
                        if 'Y' in palabras[indice_entre:]:
                            indice_y = palabras.index('Y', indice_entre)
                            palabras = palabras[:indice_y]
                        if 'Y' in palabras:
                            palabras = [palabra for palabra in palabras if palabra != 'Y']
                if 'CRUCERO' in palabras:
                    indice_entre = palabras.index('CRUCERO')
                    if indice_entre > 0 and (indice_entre + 1) < len(palabras) and palabras[indice_entre + 1] in borrar_palabras:
                        palabras[indice_entre:indice_entre + 2] = ['']
                        if 'Y' in palabras[indice_entre:]:
                            indice_y = palabras.index('Y', indice_entre)
                            palabras = palabras[:indice_y]
                        if 'Y' in palabras:
                            palabras = [palabra for palabra in palabras if palabra != 'Y']
                    elif indice_entre > 0 and (indice_entre + 2) < len(palabras) and palabras[indice_entre + 2] in borrar_palabras:
                        palabras[indice_entre:indice_entre + 3] = ['']
                        if 'Y' in palabras[indice_entre:]:
                            indice_y = palabras.index('Y', indice_entre)
                            palabras = palabras[:indice_y]
                        if 'Y' in palabras:
                            palabras = [palabra for palabra in palabras if palabra != 'Y']
                    elif indice_entre > 0 and (indice_entre + 3) < len(palabras) and palabras[indice_entre + 3] in borrar_palabras:
                        palabras[indice_entre:indice_entre + 4] = ['']
                        if 'Y' in palabras[indice_entre:]:
                            indice_y = palabras.index('Y', indice_entre)
                            palabras = palabras[:indice_y]
                        if 'Y' in palabras:
                            palabras = [palabra for palabra in palabras if palabra != 'Y']
                if 'CON' in palabras:
                    indice_entre = palabras.index('CON')
                    if indice_entre > 0 and (indice_entre + 1) < len(palabras) and palabras[indice_entre + 1] in borrar_palabras:
                        palabras[indice_entre:indice_entre + 2] = ['']
                        if 'Y' in palabras[indice_entre:]:
                            indice_y = palabras.index('Y', indice_entre)
                            palabras = palabras[:indice_y]
                        if 'Y' in palabras:
                            palabras = [palabra for palabra in palabras if palabra != 'Y']
                    elif indice_entre > 0 and (indice_entre + 2) < len(palabras) and palabras[indice_entre + 2] in borrar_palabras:
                        palabras[indice_entre:indice_entre + 3] = ['']
                        if 'Y' in palabras[indice_entre:]:
                            indice_y = palabras.index('Y', indice_entre)
                            palabras = palabras[:indice_y]
                        if 'Y' in palabras:
                            palabras = [palabra for palabra in palabras if palabra != 'Y']
                    elif indice_entre > 0 and (indice_entre + 3) < len(palabras) and palabras[indice_entre + 3] in borrar_palabras:
                        palabras[indice_entre:indice_entre + 4] = ['']
                        if 'Y' in palabras[indice_entre:]:
                            indice_y = palabras.index('Y', indice_entre)
                            palabras = palabras[:indice_y]
                        if 'Y' in palabras:
                            palabras = [palabra for palabra in palabras if palabra != 'Y']
                if 'HACIA' in palabras:
                    indice_entre = palabras.index('HACIA')
                    if indice_entre > 0 and (indice_entre + 1) < len(palabras) and palabras[indice_entre + 1] in borrar_palabras:
                        palabras[indice_entre:indice_entre + 2] = ['']
                        if 'Y' in palabras[indice_entre:]:
                            indice_y = palabras.index('Y', indice_entre)
                            palabras = palabras[:indice_y]
                        if 'Y' in palabras:
                            palabras = [palabra for palabra in palabras if palabra != 'Y']
                    elif indice_entre > 0 and (indice_entre + 2) < len(palabras) and palabras[indice_entre + 2] in borrar_palabras:
                        palabras[indice_entre:indice_entre + 3] = ['']
                        if 'Y' in palabras[indice_entre:]:
                            indice_y = palabras.index('Y', indice_entre)
                            palabras = palabras[:indice_y]
                        if 'Y' in palabras:
                            palabras = [palabra for palabra in palabras if palabra != 'Y']
                    elif indice_entre > 0 and (indice_entre + 3) < len(palabras) and palabras[indice_entre + 3] in borrar_palabras:
                        palabras[indice_entre:indice_entre + 4] = ['']
                        if 'Y' in palabras[indice_entre:]:
                            indice_y = palabras.index('Y', indice_entre)
                            palabras = palabras[:indice_y]
                        if 'Y' in palabras:
                            palabras = [palabra for palabra in palabras if palabra != 'Y']
                if 'FRENTE' in palabras:
                    indice_entre = palabras.index('FRENTE')
                    if indice_entre > 0 and (indice_entre + 1) < len(palabras) and palabras[indice_entre + 1] in borrar_palabras:
                        palabras[indice_entre:indice_entre + 2] = ['']
                        if 'Y' in palabras[indice_entre:]:
                            indice_y = palabras.index('Y', indice_entre)
                            palabras = palabras[:indice_y]
                        if 'Y' in palabras:
                            palabras = [palabra for palabra in palabras if palabra != 'Y']
                    elif indice_entre > 0 and (indice_entre + 2) < len(palabras) and palabras[indice_entre + 2] in borrar_palabras:
                        palabras[indice_entre:indice_entre + 3] = ['']
                        if 'Y' in palabras[indice_entre:]:
                            indice_y = palabras.index('Y', indice_entre)
                            palabras = palabras[:indice_y]
                        if 'Y' in palabras:
                            palabras = [palabra for palabra in palabras if palabra != 'Y']
                    elif indice_entre > 0 and (indice_entre + 3) < len(palabras) and palabras[indice_entre + 3] in borrar_palabras:
                        palabras[indice_entre:indice_entre + 4] = ['']
                        if 'Y' in palabras[indice_entre:]:
                            indice_y = palabras.index('Y', indice_entre)
                            palabras = palabras[:indice_y]
                        if 'Y' in palabras:
                            palabras = [palabra for palabra in palabras if palabra != 'Y']

                direccion_final = ' '.join(palabra for palabra in palabras if palabra)
                direcciones_estandarizadas.append(direccion_final.strip())
                #print(direcciones_estandarizadas)
            else:
                direccion_final = ' '.join(palabra for palabra in palabras if palabra)
                direcciones_estandarizadas.append(direccion_final.strip())
        nueva_direccion= ';'.join(direcciones_estandarizadas)        
        direccion_limpia.append(nueva_direccion)
    return direccion_limpia

def limpiar_direccion(direcciones):
    direccion_limpia = []
    direccion_split = []

    for direccion_larga in direcciones:
        direccion_corta = []
        encontrado = False
        direccion_split = str.upper(direccion_larga)
        direcciones = direccion_split.split(";")
        for direccion in direcciones:
            nueva_cadena = ""
            if direccion.strip() == "":
                direccion_corta.append(direccion)
            elif direccion =="NAN" or direccion=="NULL":
                direccion = ""
                direccion_corta.append(direccion)
            elif direccion.startswith("<"):
                if direccion == "<":
                    direccion = ""
                    break
                contenido = direccion[1:-1]  # Elimina los signos "<" y ">"
                split_contenido = contenido.split()
                if split_contenido[0] in palabras_clave:
                    contenido = " ".join(split_contenido)
                    if ";" in contenido:
                        parte_anterior = contenido.split(";", 1)[0]
                        dire_upper = str.upper(parte_anterior)
                        direccion_split = dire_upper.split()
                        direccion_unida= ""
                        if direccion_split[0] in palabras_clave:
                            direccion_unida = " ".join(direccion_split)
                            direccion_unida = re.compile(r'<[^>]+>')
                            cadena_sin_contenido = direccion_unida.sub('', dire_upper)
                            contenido_entre_parentesis = re.compile(r'\([^)]+\)')
                            cadena_sin_parentesis = contenido_entre_parentesis.sub('', cadena_sin_contenido)
                            for caracter in cadena_sin_parentesis:
                                if caracter in eliminar_caracteres:
                                    encontrado = True
                                if not encontrado:
                                    nueva_cadena += caracter
                            direccion_corta.append(nueva_cadena)
                    else:
                        direccion_unida = re.compile(r'<[^>]+>')
                        cadena_sin_contenido = direccion_unida.sub('', contenido)
                        contenido_entre_parentesis = re.compile(r'\([^)]+\)')
                        cadena_sin_parentesis = contenido_entre_parentesis.sub('', cadena_sin_contenido)
                        for caracter in cadena_sin_parentesis:
                            if caracter in eliminar_caracteres:
                                encontrado = True
                            if not encontrado:
                                nueva_cadena += caracter
                        direccion_corta.append(nueva_cadena)    

                elif split_contenido[0] not in palabras_clave:
                    contenido = " ".join(split_contenido)
                    if ";" in contenido:
                        parte_anterior = contenido.split(";", 1)[1]
                        dire_upper = str.upper(parte_anterior)
                        direccion_split = dire_upper.split()
                        direccion_unida= ""
                        if direccion_split[0] in palabras_clave:
                            direccion_unida = " ".join(direccion_split)
                            direccion_unida = re.compile(r'<[^>]+>')
                            cadena_sin_contenido = direccion_unida.sub('', dire_upper)
                            contenido_entre_parentesis = re.compile(r'\([^)]+\)')
                            cadena_sin_parentesis = contenido_entre_parentesis.sub('', cadena_sin_contenido)
                            for caracter in cadena_sin_parentesis:
                                if caracter in eliminar_caracteres:
                                    encontrado = True
                                if not encontrado:
                                    nueva_cadena += caracter
                            direccion_corta.append(nueva_cadena)
                        else :
                                contenido = ""
                                direccion_corta.append(contenido)  
                    else :
                        contenido = ""
                        direccion_corta.append(contenido)   

            else:
                dire_upper = str.upper(direccion)
                direccion_split = dire_upper.split()
                direccion_unida= ""
                
                if direccion_split[0] in palabras_clave:
                    direccion_unida = " ".join(direccion_split)
                    patron_separar = r'([A-Za-z0-9])<'
                    cadena_modificada = re.sub(patron_separar, r'\1 <', direccion_unida)
                    patron = r'<(.*?)>'
                    resultado = re.search(patron, direccion_unida)

                    cadena_extraida = ""
                    if resultado:
                        cadena_extraida = resultado.group(1)
                        # Encontrar el índice del primer "<" en la cadena original
                        indice_inicio = cadena_modificada.find("<")
                        indice_final = cadena_modificada.find(">")
                        # La cadena restante es desde el inicio hasta el primer "<"
                        cadena_restante = cadena_modificada[:indice_inicio]
                        cadena_restante_externa = cadena_modificada[indice_final+1:]
                    else: 
                        cadena_restante = cadena_modificada

                    # Verificar si la palabra "ENTRE" está en la cadena extraida
                    
                    cadena_extraida = cadena_extraida.split()

                    if len(cadena_extraida) == 1:
                        cadena_extraida = ' '.join(cadena_extraida)
                        cadena_completa = cadena_restante + cadena_extraida + cadena_restante_externa
                    elif cadena_extraida[0] in palabras_clave:
                        cadena_extraida=' '.join(cadena_extraida)
                        cadena_completa = cadena_restante + cadena_extraida
                    else: 
                        cadena_extraida = ' '.join(cadena_extraida)
                        if "ENTRE" in cadena_extraida:
                            
                            partes = cadena_extraida.split("ENTRE")
                            nueva_cadena_cruce = "ENTRE" + partes[1]
                        elif "POR" in cadena_extraida:
                            partes = cadena_extraida.split("POR")
                            nueva_cadena_cruce = "POR" + partes[1]
                        elif "CON" in cadena_extraida:
                            partes = cadena_extraida.split("CON")
                            nueva_cadena_cruce = "CON" + partes[1]
                        elif "CRUCERO" in cadena_extraida:
                            partes = cadena_extraida.split("CRUCERO")
                            nueva_cadena_cruce = "CRUCERO" + partes[1]
                        elif "HACIA" in cadena_extraida:
                            partes = cadena_extraida.split("HACIA")
                            nueva_cadena_cruce = "HACIA" + partes[1]
                        elif "FRENTE" in cadena_extraida:
                            partes = cadena_extraida.split("FRENTE")
                            nueva_cadena_cruce = "FRENTE" + partes[1]
                        else:
                            nueva_cadena_cruce = cadena_extraida

                        # Concatenar la cadena restante con la nueva cadena
                        cadena_completa = cadena_restante + nueva_cadena_cruce

                    if "ENTRE" in cadena_completa or "POR" in direccion_unida or "CON" in direccion_unida or "CRUCERO" in direccion_unida or "HACIA" in direccion_unida or "FRENTE" in direccion_unida:
                        cadena_nueva = cadena_completa.replace("ENTRE", "ENTRE").replace("POR", "POR").replace("CON","CON").replace("CRUCERO","CRUCERO").replace(">", "").replace("FRENTE","FRENTE").replace("HACIA","HACIA")
                        
                        direccion_sin_caracteres = re.compile(r'<[^>]+>')
                        cadena_sin_contenido = direccion_sin_caracteres.sub('', cadena_nueva)
                        contenido_entre_parentesis = re.compile(r'\([^)]+\)')
                        cadena_sin_parentesis = contenido_entre_parentesis.sub('', cadena_sin_contenido)

                        for caracter in cadena_sin_parentesis:
                            if caracter in eliminar_caracteres:
                                encontrado = True
                            if not encontrado:
                                nueva_cadena += caracter
                    else:
                        direccion_sin_caracteres = re.compile(r'<[^>]+>')
                        cadena_sin_contenido = direccion_sin_caracteres.sub('', dire_upper)
                        contenido_entre_parentesis = re.compile(r'\([^)]+\)')
                        cadena_sin_parentesis = contenido_entre_parentesis.sub('', cadena_sin_contenido)

                        for caracter in cadena_sin_parentesis:
                            if caracter in eliminar_caracteres:
                                encontrado = True
                            if not encontrado:
                                nueva_cadena += caracter
                    direccion_corta.append(nueva_cadena)
                else:
                    direccion_unida = " ".join(direccion_split)
                    direccion_sin_caracteres = re.compile(r'<[^>]+>')
                    # si dentro de  <> encuentra la palabra carrera,calle,cl...etc no me lo borre
                    # Sustituye el contenido entre '<' y '>' con una cadena vacía
                    cadena_sin_contenido = direccion_sin_caracteres.sub('', cadena_completa)
                    contenido_entre_parentesis = re.compile(r'\([^)]+\)')

                    # Sustituye el contenido entre paréntesis con una cadena vacía
                    cadena_sin_parentesis = contenido_entre_parentesis.sub('', cadena_sin_contenido)
                    
                    for caracter in cadena_sin_parentesis:
                        if caracter in eliminar_caracteres:
                            encontrado = True
                        if not encontrado:
                            nueva_cadena += caracter
                    direccion_corta.append(nueva_cadena)
        nueva_direccion= ';'.join(direccion_corta)        
        direccion_limpia.append(nueva_direccion)
    return direccion_limpia

data = pd.read_excel('paq2.xlsx')

ids = data["id"].tolist()
registros = data["Direccion"].tolist()

lista_de_registros = []

for registro in registros:
    dire_parse = str(registro)
    dire_upper = str.upper(dire_parse)
    lista_de_registros.append(dire_upper)

direccion_estandarizada = limpiar_direccion(lista_de_registros)

direccion_modificada = []

for direccion in direccion_estandarizada:
    for patron, reemplazo in sustituciones.items():
        direccion = re.sub(patron, reemplazo, direccion)
    direccion_modificada.append(direccion)

direccion_sin_cruce = borrar_entre(direccion_modificada) 
direccion_limpia = borrar_union(direccion_sin_cruce)
direccion_sin_palabras = eliminar_palabras(direccion_limpia)
direccion_sin_repetir_separada = borrar_repetido_separada(direccion_sin_palabras)
direccion_sin_repetir= borrar_repetido(direccion_sin_repetir_separada)
direccion_vacia = vaciar_invalidas(direccion_sin_repetir)

wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(row=1, column=1).value = 'ID'  # Nombre de la columna 1
sheet.cell(row=1, column=2).value = 'Dirección Estandarizada'
sheet.cell(row=1, column=3).value = 'Dirección Original' 

for i, dato in enumerate(direccion_vacia):
    sheet.cell(row=i + 2, column=1).value = ids[i]
    sheet.cell(row=i + 2, column=2).value = dato
    sheet.cell(row=i + 2, column=3).value = registros[i]

wb.save("paq2_DIRECCIONES.xlsx")